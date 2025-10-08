from flask import Flask, jsonify, request, send_file
from flask_cors import CORS
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
import os
from datetime import datetime
from database import (
    init_database, obtener_subastas, obtener_imagenes_subasta,
    obtener_documentos_subasta, get_db_connection
)
from scraper import scraping_completo
import threading

app = Flask(__name__)
CORS(app)

# Inicializar base de datos al arrancar
init_database()

@app.route('/')
def home():
    return jsonify({
        "message": "Auction Brokers API",
        "status": "running",
        "version": "3.0",
        "endpoints": [
            "/api/health",
            "/api/subastas",
            "/api/subasta/<id>",
            "/api/exportar",
            "/api/stats",
            "/api/scraping/iniciar"
        ]
    })

@app.route('/api/health')
def health():
    try:
        conn = get_db_connection()
        cur = conn.cursor()
        cur.execute('SELECT COUNT(*) as total FROM subastas')
        result = cur.fetchone()
        total = result['total'] if result else 0
        cur.close()
        conn.close()
        
        return jsonify({
            "success": True,
            "status": "running",
            "database": "connected",
            "total_subastas": total
        })
    except Exception as e:
        return jsonify({
            "success": False,
            "status": "error",
            "error": str(e)
        }), 500

@app.route('/api/subastas')
def get_subastas():
    try:
        provincia = request.args.get('provincia', '')
        tipo = request.args.get('tipo', '')
        search = request.args.get('search', '')
        
        filtros = {}
        if provincia:
            filtros['provincia'] = provincia
        if tipo:
            filtros['tipo_bien'] = tipo
        if search:
            filtros['search'] = search
        
        resultados = obtener_subastas(filtros)
        
        # Convertir resultados a formato JSON serializable
        subastas = []
        for subasta in resultados:
            subasta_dict = dict(subasta)
            
            # Convertir fechas a string
            if subasta_dict.get('fecha_inicio'):
                subasta_dict['fecha_inicio'] = subasta_dict['fecha_inicio'].isoformat()
            if subasta_dict.get('fecha_conclusion'):
                subasta_dict['fecha_conclusion'] = subasta_dict['fecha_conclusion'].isoformat()
            if subasta_dict.get('fecha_scraping'):
                subasta_dict['fecha_scraping'] = subasta_dict['fecha_scraping'].isoformat()
            if subasta_dict.get('actualizado'):
                subasta_dict['actualizado'] = subasta_dict['actualizado'].isoformat()
            
            # Convertir Decimal a float
            for key in ['cantidad_reclamada', 'valor_tasacion', 'valor_subasta', 
                       'tramos_pujas', 'puja_minima', 'puja_maxima', 'importe_deposito',
                       'latitud', 'longitud']:
                if subasta_dict.get(key) is not None:
                    subasta_dict[key] = float(subasta_dict[key])
            
            # Agregar coordenadas en formato esperado por el frontend
            if subasta_dict.get('latitud') and subasta_dict.get('longitud'):
                subasta_dict['coordenadas'] = {
                    'lat': float(subasta_dict['latitud']),
                    'lng': float(subasta_dict['longitud'])
                }
            
            # Obtener imágenes y documentos
            imagenes = obtener_imagenes_subasta(subasta_dict['id'])
            documentos = obtener_documentos_subasta(subasta_dict['id'])
            
            subasta_dict['imagenes'] = [
                {
                    'nombre': img['nombre'],
                    'url': img['url_s3'] or img['url_original']
                }
                for img in imagenes
            ]
            
            subasta_dict['documentos'] = [
                {
                    'nombre': doc['nombre'],
                    'url': doc['url_s3'] or doc['url_original'],
                    'size': f"{doc['size_bytes'] / 1024:.0f} KB" if doc['size_bytes'] else "N/A"
                }
                for doc in documentos
            ]
            
            subastas.append(subasta_dict)
        
        return jsonify({
            "success": True,
            "data": subastas,
            "total": len(subastas)
        })
        
    except Exception as e:
        return jsonify({
            "success": False,
            "error": str(e)
        }), 500

@app.route('/api/subasta/<subasta_id>')
def get_subasta_detalle(subasta_id):
    try:
        conn = get_db_connection()
        cur = conn.cursor()
        
        cur.execute('SELECT * FROM subastas WHERE id = %s', (subasta_id,))
        subasta = cur.fetchone()
        
        if not subasta:
            return jsonify({"success": False, "error": "Subasta no encontrada"}), 404
        
        # Convertir a dict y preparar datos
        subasta_dict = dict(subasta)
        
        # Convertir fechas
        if subasta_dict.get('fecha_inicio'):
            subasta_dict['fecha_inicio'] = subasta_dict['fecha_inicio'].isoformat()
        if subasta_dict.get('fecha_conclusion'):
            subasta_dict['fecha_conclusion'] = subasta_dict['fecha_conclusion'].isoformat()
        
        # Obtener imágenes y documentos
        imagenes = obtener_imagenes_subasta(subasta_id)
        documentos = obtener_documentos_subasta(subasta_id)
        
        subasta_dict['imagenes'] = [dict(img) for img in imagenes]
        subasta_dict['documentos'] = [dict(doc) for doc in documentos]
        
        cur.close()
        conn.close()
        
        return jsonify({"success": True, "data": subasta_dict})
        
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500

@app.route('/api/stats')
def get_stats():
    try:
        conn = get_db_connection()
        cur = conn.cursor()
        
        # Total subastas
        cur.execute('SELECT COUNT(*) as total FROM subastas')
        total = cur.fetchone()['total']
        
        # Por provincia
        cur.execute('''
            SELECT provincia, COUNT(*) as cantidad
            FROM subastas
            GROUP BY provincia
            ORDER BY cantidad DESC
            LIMIT 10
        ''')
        por_provincia = [dict(row) for row in cur.fetchall()]
        
        # Por tipo de bien
        cur.execute('''
            SELECT tipo_bien, COUNT(*) as cantidad
            FROM subastas
            GROUP BY tipo_bien
            ORDER BY cantidad DESC
        ''')
        por_tipo = [dict(row) for row in cur.fetchall()]
        
        # Por estado
        cur.execute('''
            SELECT estado, COUNT(*) as cantidad
            FROM subastas
            GROUP BY estado
            ORDER BY cantidad DESC
        ''')
        por_estado = [dict(row) for row in cur.fetchall()]
        
        cur.close()
        conn.close()
        
        return jsonify({
            "success": True,
            "stats": {
                "total": total,
                "por_provincia": por_provincia,
                "por_tipo": por_tipo,
                "por_estado": por_estado
            }
        })
        
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500

@app.route('/api/exportar', methods=['POST'])
def exportar_excel():
    try:
        data = request.get_json() or {}
        ids = data.get('ids', [])
        
        if not ids:
            # Exportar todas
            subastas_exportar = obtener_subastas()
        else:
            # Exportar seleccionadas
            conn = get_db_connection()
            cur = conn.cursor()
            placeholders = ','.join(['%s'] * len(ids))
            cur.execute(f'SELECT * FROM subastas WHERE id IN ({placeholders})', ids)
            subastas_exportar = cur.fetchall()
            cur.close()
            conn.close()
        
        # Crear Excel
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Subastas BOE"
        
        # Estilos
        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        # Encabezados
        headers = [
            "RATIO 1 - cantidad reclamada vs valor subasta",
            "RATIO 2 - Puja Máxima vs Valor Subasta",
            "Estado",
            "Tipo De Subasta",
            "Tipo Bien",
            "Id",
            "Lotes",
            "Provincia",
            "Localidad",
            "Dirección",
            "Boton google maps",
            "Descripción",
            "Referencia Catastral",
            "Marca",
            "Modelo",
            "Matricula",
            "Cantidad Reclamada",
            "Valor De Tasacion",
            "Valor Subasta",
            "Tramos Entre Pujas",
            "Puja Mínima",
            "Puja Máxima",
            "Importe Del Deposito",
            "Nombre",
            "Fecha De Inicio",
            "Fecha De Conclusión"
        ]
        
        # Escribir encabezados
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
        
        # Escribir datos
        for row_num, subasta in enumerate(subastas_exportar, 2):
            subasta_dict = dict(subasta)
            
            google_maps_url = ""
            if subasta_dict.get('latitud') and subasta_dict.get('longitud'):
                lat = float(subasta_dict['latitud'])
                lng = float(subasta_dict['longitud'])
                google_maps_url = f"https://www.google.com/maps/search/?api=1&query={lat},{lng}"
            
            # RATIO 1
            ws.cell(row=row_num, column=1).value = f"=IF(Q{row_num}=0,0,(Q{row_num}/S{row_num})*100)"
            ws.cell(row=row_num, column=1).number_format = '0.00"%"'
            
            # RATIO 2
            ws.cell(row=row_num, column=2).value = f"=IF(V{row_num}=0,0,(V{row_num}/S{row_num})*100)"
            ws.cell(row=row_num, column=2).number_format = '0.00"%"'
            
            # Resto de columnas
            ws.cell(row=row_num, column=3).value = subasta_dict.get('estado', '')
            ws.cell(row=row_num, column=4).value = subasta_dict.get('tipo_subasta', '')
            ws.cell(row=row_num, column=5).value = subasta_dict.get('tipo_bien', '')
            ws.cell(row=row_num, column=6).value = subasta_dict.get('id', '')
            ws.cell(row=row_num, column=7).value = subasta_dict.get('lotes', '')
            ws.cell(row=row_num, column=8).value = subasta_dict.get('provincia', '')
            ws.cell(row=row_num, column=9).value = subasta_dict.get('localidad', '')
            ws.cell(row=row_num, column=10).value = subasta_dict.get('direccion', '')
            
            if google_maps_url:
                cell = ws.cell(row=row_num, column=11)
                cell.value = "Ver en Google Maps"
                cell.hyperlink = google_maps_url
                cell.font = Font(color="0563C1", underline="single")
            
            ws.cell(row=row_num, column=12).value = subasta_dict.get('descripcion', '')
            ws.cell(row=row_num, column=13).value = subasta_dict.get('referencia_catastral', '')
            ws.cell(row=row_num, column=14).value = subasta_dict.get('marca', '')
            ws.cell(row=row_num, column=15).value = subasta_dict.get('modelo', '')
            ws.cell(row=row_num, column=16).value = subasta_dict.get('matricula', '')
            ws.cell(row=row_num, column=17).value = float(subasta_dict.get('cantidad_reclamada', 0) or 0)
            ws.cell(row=row_num, column=18).value = float(subasta_dict.get('valor_tasacion', 0) or 0)
            ws.cell(row=row_num, column=19).value = float(subasta_dict.get('valor_subasta', 0) or 0)
            ws.cell(row=row_num, column=20).value = float(subasta_dict.get('tramos_pujas', 0) or 0)
            ws.cell(row=row_num, column=21).value = float(subasta_dict.get('puja_minima', 0) or 0)
            ws.cell(row=row_num, column=22).value = float(subasta_dict.get('puja_maxima', 0) or 0)
            ws.cell(row=row_num, column=23).value = float(subasta_dict.get('importe_deposito', 0) or 0)
            ws.cell(row=row_num, column=24).value = subasta_dict.get('nombre_acreedor', '')
            
            # Fechas
            fecha_inicio = subasta_dict.get('fecha_inicio')
            if fecha_inicio:
                ws.cell(row=row_num, column=25).value = fecha_inicio.strftime('%d/%m/%Y') if hasattr(fecha_inicio, 'strftime') else str(fecha_inicio)
            
            fecha_conclusion = subasta_dict.get('fecha_conclusion')
            if fecha_conclusion:
                ws.cell(row=row_num, column=26).value = fecha_conclusion.strftime('%d/%m/%Y') if hasattr(fecha_conclusion, 'strftime') else str(fecha_conclusion)
        
        # Ajustar anchos de columna
        column_widths = [35, 35, 12, 15, 12, 18, 15, 12, 15, 40, 20, 50, 25, 15, 15, 12, 18, 18, 15, 18, 15, 15, 18, 40, 15, 18]
        for col_num, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(col_num)].width = width
        
        # Congelar primera fila
        ws.freeze_panes = "A2"
        
        # Añadir filtros
        ws.auto_filter.ref = ws.dimensions
        
        # Guardar archivo temporal
        os.makedirs('temp', exist_ok=True)
        filename = f'temp/subastas_boe_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        wb.save(filename)
        
        return send_file(
            filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'subastas_boe_{datetime.now().strftime("%Y%m%d")}.xlsx'
        )
    
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/scraping/iniciar', methods=['POST'])
def iniciar_scraping():
    """Iniciar scraping en segundo plano"""
    try:
        # Ejecutar scraping en un thread separado
        thread = threading.Thread(target=scraping_completo)
        thread.daemon = True
        thread.start()
        
        return jsonify({
            "success": True,
            "message": "Scraping iniciado en segundo plano. Puede tardar varias horas."
        })
    except Exception as e:
        return jsonify({
            "success": False,
            "error": str(e)
        }), 500

if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5000))
    app.run(host='0.0.0.0', port=port)
